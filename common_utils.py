# -*- coding: utf-8 -*-
"""
공통 유틸리티 함수들
"""
import os
import requests
import pandas as pd
from datetime import datetime, timedelta, timezone
import re
import time
from dotenv import load_dotenv
from pathlib import Path
from google.oauth2 import service_account
from googleapiclient.discovery import build

# .env 파일 로드
load_dotenv()

# 기본 설정 - GitHub Actions에서는 임시 디렉토리 사용
if os.getenv('GITHUB_ACTIONS'):
    # GitHub Actions 환경에서는 임시 디렉토리 사용 (업로드하지 않음)
    DOWNLOAD_DIR = "/tmp/3pl_temp"
    print("🔒 GitHub Actions 환경: 임시 디렉토리 사용 (보안)")
else:
    # 로컬 환경에서는 기존 경로 사용
    DOWNLOAD_DIR = os.path.join(Path.home(), "Documents", "3pl", "daily")

os.makedirs(DOWNLOAD_DIR, exist_ok=True)

def get_korean_holidays(year):
    """Nager.Date API를 통해 한국 공휴일 정보 가져오기"""
    try:
        import requests
        url = f'https://date.nager.at/api/v3/PublicHolidays/{year}/KR'
        response = requests.get(url, timeout=10)
        
        if response.status_code == 200:
            holidays = response.json()
            holiday_dates = {holiday['date'] for holiday in holidays}
            print(f"📅 {year}년 한국 공휴일 {len(holiday_dates)}개 로드됨")
            return holiday_dates
        else:
            print(f"⚠️ 공휴일 API 오류 (상태: {response.status_code}), 하드코딩된 공휴일 사용")
            return get_fallback_holidays(year)
            
    except Exception as e:
        print(f"⚠️ 공휴일 API 연결 실패: {e}, 하드코딩된 공휴일 사용")
        return get_fallback_holidays(year)

def get_fallback_holidays(year):
    """API 실패 시 사용할 하드코딩된 공휴일 (2025년 기준)"""
    if year == 2025:
        return {
            "2025-01-01", "2025-01-28", "2025-01-29", "2025-01-30",
            "2025-03-01", "2025-05-01", "2025-05-05", "2025-05-06",
            "2025-06-03", "2025-06-06", "2025-08-15",
            "2025-10-03",
            "2025-10-06", "2025-10-07", "2025-10-08", "2025-10-09", "2025-12-25"
        }
    else:
        # 다른 연도는 기본 공휴일만 포함
        return {
            f"{year}-01-01",  # 신정
            f"{year}-03-01",  # 삼일절
            f"{year}-05-05",  # 어린이날
            f"{year}-06-06",  # 현충일
            f"{year}-08-15",  # 광복절
            f"{year}-10-03",  # 개천절
            f"{year}-10-09",  # 한글날
            f"{year}-12-25"   # 크리스마스
        }

def is_holiday(date):
    """공휴일 확인 (API 기반)"""
    year = date.year
    holidays = get_korean_holidays(year)
    return date.strftime("%Y-%m-%d") in holidays

def get_last_business_day(date):
    """마지막 영업일 계산"""
    while date.weekday() >= 5 or is_holiday(date):
        date -= timedelta(days=1)
    return date

def should_skip_today():
    """오늘 작업을 건너뛸지 확인 (공휴일/주말 체크)"""
    # KST(한국 표준시) = UTC+9
    kst_offset = timezone(timedelta(hours=9))
    today_kst = datetime.now(kst_offset).date()
    
    # 요일 확인 (0=월, 6=일)
    weekday = today_kst.weekday()
    weekday_names = ['월', '화', '수', '목', '금', '토', '일']
    is_weekend = weekday >= 5  # 토(5), 일(6)
    
    # 공휴일 확인
    is_holiday_today = is_holiday(today_kst)
    
    print(f"📅 오늘 날짜: {today_kst} ({weekday_names[weekday]}요일)")
    
    if is_holiday_today:
        print(f"🎉 오늘은 공휴일입니다.")
    
    if is_weekend:
        print(f"🏖️ 오늘은 주말입니다.")
    
    # 특정 휴무일 확인
    is_custom_holiday_today = is_custom_holiday(today_kst)
    
    if is_custom_holiday_today:
        print(f"🚫 오늘은 지정된 휴무일입니다.")
    
    should_skip = is_holiday_today or is_weekend or is_custom_holiday_today
    
    if should_skip:
        skip_reason = []
        if is_holiday_today:
            skip_reason.append("공휴일")
        if is_weekend:
            skip_reason.append("주말")
        if is_custom_holiday_today:
            skip_reason.append("지정휴무일")
        
        print(f"⏭️ 작업 건너뛰기: {', '.join(skip_reason)}")
        return True
    else:
        print(f"✅ 작업 진행 가능: 평일")
        return False

def is_custom_holiday(date_obj):
    """특정 휴무일인지 확인 (코드 내 지정)"""
    
    # 코드 내 특정 휴무일 지정 (YYYY-MM-DD 형식)
    CUSTOM_HOLIDAYS = [
        "2025-10-02",  # 추석 전 배송 안 함
        # 예시: "2025-01-01",  # 신정  
        # 예시: "2025-08-15",  # 임시 휴무
        # 필요시 여기에 날짜 추가
    ]
    
    date_str = date_obj.strftime('%Y-%m-%d')
    is_custom = date_str in CUSTOM_HOLIDAYS
    
    if is_custom:
        print(f"🔍 특정 휴무일 감지: {date_str}")
    
    return is_custom

def find_last_work_day(current_date):
    """마지막 작업일 찾기 (연속 휴무일 역추적)"""
    
    check_date = current_date - timedelta(days=1)  # 어제부터 시작
    days_checked = 0
    
    while days_checked < 14:  # 최대 2주까지만 역추적 (무한루프 방지)
        # 주말인지 확인
        if check_date.weekday() >= 5:
            check_date -= timedelta(days=1)
            days_checked += 1
            continue
        
        # 공휴일인지 확인
        if is_holiday(check_date):
            check_date -= timedelta(days=1)
            days_checked += 1
            continue
        
        # 특정 휴무일인지 확인
        if is_custom_holiday(check_date):
            check_date -= timedelta(days=1)
            days_checked += 1
            continue
        
        # 작업일 발견
        return check_date
    
    # 최대 역추적 한계에 도달한 경우 기본값 반환
    return current_date - timedelta(days=1)

def get_date_range():
    """주문 조회 날짜 범위 계산 (연속 휴무일 고려)"""
    # KST(한국 표준시) = UTC+9
    kst_offset = timezone(timedelta(hours=9))
    today_kst = datetime.now(kst_offset)
    
    print(f"📅 오늘 날짜: {today_kst.strftime('%Y-%m-%d')} ({['월요일', '화요일', '수요일', '목요일', '금요일', '토요일', '일요일'][today_kst.weekday()]})")
    
    # 마지막 작업일 찾기 (연속 휴무일 역추적)
    last_work_day = find_last_work_day(today_kst)
    
    print(f"📅 마지막 작업일: {last_work_day.strftime('%Y-%m-%d')} ({['월요일', '화요일', '수요일', '목요일', '금요일', '토요일', '일요일'][last_work_day.weekday()]})")
    
    # 연속 휴무일 계산
    days_gap = (today_kst.date() - last_work_day.date()).days
    
    if days_gap > 1:
        print(f"🔍 연속 휴무: {days_gap}일간 휴무 감지")
        print(f"📈 확장 처리: {last_work_day.strftime('%Y-%m-%d')} 12시 ~ {today_kst.strftime('%Y-%m-%d')} 12시")
        
        # 확장된 범위: 마지막 작업일 12시 ~ 오늘 12시
        start_date = last_work_day.replace(hour=12, minute=0, second=0, microsecond=0, tzinfo=kst_offset)
        end_date = today_kst.replace(hour=12, minute=0, second=0, microsecond=0)
        
        print(f"📅 확장 조회 기간: {start_date} ~ {end_date}")
        print(f"📅 처리 기간: {days_gap}일 + {(end_date - start_date).seconds // 3600}시간")
        
        return start_date, end_date
    else:
        # 일반 처리 (1일 차이)
        print("📅 일반 처리: 전날 ~ 당일")
        
        yesterday = today_kst.date() - timedelta(days=1)
        start_date = datetime.combine(yesterday, datetime.min.time()).replace(hour=12, tzinfo=kst_offset)
        end_date = today_kst.replace(hour=12, minute=0, second=0, microsecond=0)
        
        weekday_name = ['월', '화', '수', '목', '금', '토', '일'][today_kst.weekday()]
        print(f"📅 {weekday_name}요일 일반 처리: 전날 ~ 당일")
        print(f"📅 주문 조회 시간 (KST): {start_date.strftime('%Y-%m-%d %H:%M %z')} ~ {end_date.strftime('%Y-%m-%d %H:%M %z')}")
        print(f"📅 처리 기간: 24시간")
        
        return start_date, end_date

def get_woocommerce_auth(base_url, consumer_key, consumer_secret):
    """WooCommerce Consumer Key/Secret 인증 설정 및 테스트"""
    print(f"🔍 WooCommerce Consumer Key/Secret 인증 준비: {base_url}")
    if consumer_key and consumer_secret:
        print(f"✅ WooCommerce 인증 정보 준비 완료: {base_url}")
        print(f"🔍 Consumer Key: {consumer_key[:10]}...")
        
        # 간단한 연결 테스트
        test_url = f"{base_url}/wp-json/wc/v3/system_status"
        
        if base_url.startswith('https://'):
            test_params = {
                'consumer_key': consumer_key,
                'consumer_secret': consumer_secret
            }
            test_auth = None
        else:
            test_params = {}
            test_auth = (consumer_key, consumer_secret)
        
        try:
            test_response = requests.get(test_url, auth=test_auth, params=test_params, timeout=10)
            print(f"🔍 연결 테스트 응답: {test_response.status_code}")
            if test_response.status_code == 200:
                print(f"✅ WooCommerce API 연결 성공: {base_url}")
            else:
                print(f"❌ WooCommerce API 연결 실패: {test_response.status_code}")
        except Exception as e:
            print(f"❌ WooCommerce API 테스트 오류: {e}")
        
        return (consumer_key, consumer_secret)
    else:
        print(f"❌ WooCommerce Consumer Key 또는 Secret이 없습니다: {base_url}")
        return None

def fetch_orders_from_wp(base_url, auth_info, start_date, end_date, status='completed'):
    """WooCommerce REST API를 통한 주문 데이터 수집"""
    orders_url = f"{base_url}/wp-json/wc/v3/orders"
    
    print(f"🔍 WooCommerce API 호출: {base_url}")
    
    consumer_key, consumer_secret = auth_info
    
    # HTTPS 사이트의 경우 URL 파라미터로 인증 정보 전달
    if base_url.startswith('https://'):
        print(f"🔍 HTTPS 사이트 - URL 파라미터 방식 사용")
        auth = None
        headers = {'Content-Type': 'application/json'}
        
        # KST 시간을 ISO 형식으로 변환 (UTC 오프셋 포함)
        start_iso = start_date.isoformat()
        end_iso = end_date.isoformat()
        
        print(f"🔍 조회 기간 (ISO): {start_iso} ~ {end_iso}")
        
        # 날짜 범위 파라미터 + 인증 파라미터
        params = {
            'consumer_key': consumer_key,
            'consumer_secret': consumer_secret,
            'after': start_iso,
            'before': end_iso,
            'status': status,
            'per_page': 100,
            'page': 1
        }
    else:
        print(f"🔍 HTTP 사이트 - Basic Auth 방식 사용")
        auth = auth_info
        headers = {'Content-Type': 'application/json'}
        
        # KST 시간을 ISO 형식으로 변환
        start_iso = start_date.isoformat()
        end_iso = end_date.isoformat()
        
        # 날짜 범위 파라미터만
        params = {
            'after': start_iso,
            'before': end_iso,
            'status': status,
            'per_page': 100,
            'page': 1
        }
    
    all_orders = []
    
    try:
        while True:
            response = requests.get(orders_url, auth=auth, headers=headers, params=params)
            
            print(f"🔍 API 응답 상태: {response.status_code}")
            if response.status_code != 200:
                print(f"❌ 주문 데이터 수집 실패: {response.status_code}")
                print(f"❌ 응답 내용: {response.text[:500]}")
                break
                
            orders = response.json()
            if not orders:
                break
                
            all_orders.extend(orders)
            params['page'] += 1
            
            # 페이지네이션 확인
            if len(orders) < params['per_page']:
                break
                
        print(f"✅ 주문 데이터 수집 완료: {len(all_orders)}개")
        return all_orders
        
    except Exception as e:
        print(f"❌ 주문 데이터 수집 오류: {e}")
        return []

def convert_orders_to_dataframe(orders, site_label):
    """주문 데이터를 DataFrame으로 변환 (상품명 매핑 적용)"""
    
    # Google Sheets에서 상품명 매핑 데이터 가져오기
    product_mapping = get_product_name_mapping()
    
    order_items = []
    
    for order in orders:
        order_id = order.get('id', '')
        order_status = order.get('status', '')
        
        # 배송 정보
        shipping = order.get('shipping', {})
        billing = order.get('billing', {})
        
        # 고객 정보
        customer_note = order.get('customer_note', '')
        
        # 주문 상품들
        line_items = order.get('line_items', [])
        
        for item in line_items:
            # SKU 처리 (원본 보존 + 매핑용 정리)
            raw_sku = item.get('sku', '')
            
            # 1. 하이픈 이하 제거 (상태 판별용 - 디지털/B2B/예약상품 정보 보존)
            sku_for_status = raw_sku.split('-')[0] if raw_sku else ''
            
            # 2. 복합 SKU 처리: 전체 SKU에서 [디지털], [B2B], [예약상품] 체크
            # 복합 SKU라도 전체를 하나의 상품으로 처리 (개별 분리하지 않음)
            
            # 3. 상품명 매핑용 clean_sku (대괄호까지 제거, 복합 SKU도 전체 유지)
            import re
            clean_sku = re.sub(r'\[.*?\]', '', sku_for_status).strip()
            
            # Google Sheets에서 매핑된 상품명 가져오기 (없으면 원래 상품명 사용)
            original_product_name = item.get('name', '')
            mapped_product_name = product_mapping.get(clean_sku, original_product_name)
            
            # 매핑 결과 간단 로그 (처음 3개만)
            if len(order_items) < 3:
                if mapped_product_name != original_product_name:
                    print(f"   📋 상품명 매핑: {clean_sku} → {mapped_product_name}")
                else:
                    print(f"   📋 매핑 없음: {clean_sku} → 원래 상품명 사용")
            
            order_items.append({
                '주문번호': str(order_id),
                '주문상태': '완료됨' if order_status == 'completed' else order_status,
                'SKU': sku_for_status,  # 상태 판별용 SKU (디지털/B2B/예약상품 정보 보존)
                '상품명': mapped_product_name,  # 매핑된 상품명 사용
                '품번코드': clean_sku,  # 매핑용 clean SKU (대괄호 제거)
                '쇼핑몰상품코드': raw_sku,  # 원래 SKU 값 (하이픈 포함)
                '수량': str(item.get('quantity', 1)),
                '수령인명': (shipping.get('first_name', '') + ' ' + shipping.get('last_name', '')).strip(),
                '수령인 연락처': billing.get('phone', ''),
                '수령인 이메일': billing.get('email', ''),  # EMS용 이메일 정보 추가
                '우편번호': shipping.get('postcode', ''),
                '배송지주소': build_clean_address(shipping),
                '배송메세지': customer_note
            })
    
    df = pd.DataFrame(order_items)
    print(f"✅ {site_label} 주문 데이터 변환 완료: {len(df)}개 항목 (상품명 매핑 적용)")
    return df

def is_korean_address(addr):
    """한국어 주소 확인 (한글 포함 또는 KR 국가코드)"""
    if pd.isna(addr):
        return False
    
    addr_str = str(addr)
    
    # 1. 한글이 포함된 경우
    if re.search(r'[가-힣]', addr_str):
        return True
    
    # 2. 국가코드가 KR인 경우
    if re.search(r'\bKR\b', addr_str, re.IGNORECASE):
        return True
    
    # 3. 한국 관련 키워드가 포함된 경우
    korean_keywords = ['KOREA', 'SOUTH KOREA', '대한민국', '한국']
    for keyword in korean_keywords:
        if keyword in addr_str.upper():
            return True
    
    return False

def has_korean_characters(text):
    """텍스트에 한글이 포함되어 있는지 확인"""
    if pd.isna(text):
        return False
    return bool(re.search(r'[가-힣]', str(text)))

def has_non_english_characters(text):
    """텍스트에 영문이 아닌 문자(일본어, 중국어 등)가 포함되어 있는지 확인"""
    if pd.isna(text):
        return False
    
    text_str = str(text)
    
    # 1. 한글 (가-힣)
    if re.search(r'[가-힣]', text_str):
        return True
    
    # 2. 중국어/일본어 한자 (CJK Unified Ideographs)
    if re.search(r'[\u4e00-\u9fff]', text_str):
        return True
    
    # 3. 일본어 히라가나 (ひらがな)
    if re.search(r'[\u3040-\u309f]', text_str):
        return True
    
    # 4. 일본어 가타카나 (カタカナ)
    if re.search(r'[\u30a0-\u30ff]', text_str):
        return True
    
    # 5. 기타 비라틴 문자 (아랍어, 키릴문자 등)
    # 라틴 문자(확장 포함), 숫자, 기본 구두점, 공백을 제외한 나머지
    # 라틴 확장 문자 포함 (독일어 ß, 프랑스어 é, ç 등 허용)
    latin_extended_and_common = re.sub(r'[a-zA-Z0-9\s\.,\-\(\)\/\#\&\'\"\u00C0-\u00FF\u0100-\u017F]', '', text_str)
    if latin_extended_and_common.strip():
        return True
    
    return False

def filter_korean_recipients(df):
    """EMS 발송에서 한글 수령인명을 가진 주문들을 필터링하여 분리"""
    if df.empty:
        return df, pd.DataFrame()
    
    # 한글 수령인명 확인
    korean_recipients_mask = df['수령인명'].apply(has_korean_characters)
    
    # 한글 이름과 영문 이름 분리
    korean_orders = df[korean_recipients_mask].copy()
    valid_orders = df[~korean_recipients_mask].copy()
    
    if not korean_orders.empty:
        print(f"⚠️ 한글 수령인명으로 인해 EMS 처리에서 제외된 주문: {len(korean_orders)}개")
        for idx, row in korean_orders.iterrows():
            print(f"   - 주문번호 {row['주문번호']}: '{row['수령인명']}' (주소: {str(row['배송지주소'])[:50]}...)")
    
    return valid_orders, korean_orders

def filter_non_english_addresses(df):
    """EMS 발송에서 비영문 주소를 가진 주문들을 필터링하여 분리"""
    if df.empty:
        return df, pd.DataFrame()
    
    # 비영문 주소 확인
    non_english_address_mask = df['배송지주소'].apply(has_non_english_characters)
    
    # 비영문 주소와 영문 주소 분리
    non_english_orders = df[non_english_address_mask].copy()
    valid_orders = df[~non_english_address_mask].copy()
    
    if not non_english_orders.empty:
        print(f"⚠️ 비영문 주소로 인해 EMS 처리에서 제외된 주문: {len(non_english_orders)}개")
        for idx, row in non_english_orders.iterrows():
            address_preview = str(row['배송지주소'])[:50] + "..." if len(str(row['배송지주소'])) > 50 else str(row['배송지주소'])
            print(f"   - 주문번호 {row['주문번호']}: '{row['수령인명']}' (주소: {address_preview})")
    
    return valid_orders, non_english_orders

def is_pure_digital_product(sku):
    """순수 디지털 상품인지 판별 (실물 패키지 + 디지털 보너스와 구분)"""
    if pd.isna(sku):
        return False
    
    sku_str = str(sku)
    
    # SKU가 [디지털]로 끝나는지 확인
    if not sku_str.endswith('[디지털]'):
        return False
    
    # SKU 구성요소 분석
    if '/' in sku_str:
        # 복합 상품인 경우 - 실물 구성요소가 많으면 실물 패키지로 판단
        components = sku_str.split('/')
        physical_components = [c for c in components if not c.endswith('[디지털]')]
        
        # 실물 구성요소가 3개 이상이면 실물 패키지로 판단
        if len(physical_components) >= 3:
            print(f"🎯 실물 패키지 + 디지털 보너스로 판단: {sku_str[:50]}...")
            return False
    
    # 단일 상품이거나 실물 구성요소가 적으면 순수 디지털
    return True

def build_clean_address(shipping):
    """WooCommerce 배송 정보에서 중복 없는 깔끔한 주소 생성"""
    if not shipping:
        return ""
    
    # 각 필드 추출
    address_1 = shipping.get('address_1', '').strip()
    address_2 = shipping.get('address_2', '').strip()
    city = shipping.get('city', '').strip()
    state = shipping.get('state', '').strip()
    country = shipping.get('country', '').strip()
    
    # 주소 구성요소들 수집
    address_parts = []
    
    # 1. 기본 주소 (address_1)
    if address_1:
        address_parts.append(address_1)
    
    # 2. 상세 주소 (address_2)
    if address_2:
        address_parts.append(address_2)
    
    # 3. 지역 정보 중복 제거 처리 (개선된 로직)
    region_parts = []
    
    # 모든 지역 정보를 수집한 후 중복 제거
    potential_regions = []
    if state:
        potential_regions.append(state)
    if city and city != state:
        potential_regions.append(city)
    if (country and 
        country.upper() not in ['KR', 'KOREA', 'SOUTH KOREA', '대한민국', '한국']):
        potential_regions.append(country)
    
    # 중복 제거: 한 지역이 다른 지역에 포함되어 있으면 제거
    for region in potential_regions:
        is_duplicate = False
        for other_region in potential_regions:
            if region != other_region and region in other_region:
                # region이 other_region에 포함되어 있으면 중복으로 간주
                is_duplicate = True
                break
        if not is_duplicate:
            region_parts.append(region)
    
    # 주소 구성 (한국 vs 해외 구분)
    is_korean = (state and any(keyword in state for keyword in ['도', '시', '특별시', '광역시']) or
                 country and country.upper() in ['KR', 'KOREA', 'SOUTH KOREA', '대한민국', '한국'])
    
    if is_korean:
        # 한국 주소: 큰 단위 → 작은 단위 (시도 → 시군구 → 상세주소)
        final_parts = region_parts + address_parts
    else:
        # 해외 주소: 작은 단위 → 큰 단위 (상세주소 → 시 → 주/도 → 국가)
        final_parts = address_parts + region_parts
    
    # 최종 주소 생성
    full_address = ' '.join(final_parts).strip()
    
    # 연속된 공백 정리 및 추가 중복 제거
    full_address = re.sub(r'\s+', ' ', full_address)
    
    # 같은 단어가 연속으로 나오는 경우 제거 (예: "서울시 서울시" → "서울시")
    words = full_address.split()
    cleaned_words = []
    for word in words:
        if not cleaned_words or word != cleaned_words[-1]:
            cleaned_words.append(word)
    
    return ' '.join(cleaned_words)

def clean_korean_address(addr):
    """한국 주소에서 불필요한 'KR' 제거"""
    if pd.isna(addr):
        return addr
    
    # 'KR', 'KOREA', 'South Korea' 등 제거
    cleaned = str(addr)
    patterns_to_remove = [
        r'\bKR\b',
        r'\bKOREA\b', 
        r'\bSOUTH KOREA\b',
        r'\b대한민국\b',
        r'\b한국\b'
    ]
    
    for pattern in patterns_to_remove:
        cleaned = re.sub(pattern, '', cleaned, flags=re.IGNORECASE)
    
    # 연속된 공백과 콤마 정리
    cleaned = re.sub(r'\s*,\s*,\s*', ', ', cleaned)
    cleaned = re.sub(r'\s+', ' ', cleaned)
    cleaned = cleaned.strip(' ,')
    
    return cleaned

def apply_string_format(filepath, columns):
    """Excel 파일의 특정 컬럼을 문자열 형식으로 변경"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("❌ openpyxl 패키지가 설치되지 않았습니다. pip install openpyxl을 실행해주세요.")
        return
    
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        col_idx = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
        
        for col in columns:
            if col in col_idx:
                for row in ws.iter_rows(min_row=2, min_col=col_idx[col], max_col=col_idx[col]):
                    for cell in row:
                        cell.number_format = '@'
        
        wb.save(filepath)
        print(f"✅ Excel 형식 적용 완료: {filepath}")
    except Exception as e:
        print(f"❌ Excel 형식 적용 실패: {e}")

def get_product_name_mapping():
    """Supabase에서 품번코드-상품명 매핑 데이터 가져오기"""
    
    print("📋 상품명 매핑 데이터 가져오는 중 (Supabase)...")
    
    try:
        import requests
        
        # Supabase 환경변수
        supabase_url = os.getenv('SUPABASE_URL')
        supabase_key = os.getenv('SUPABASE_KEY')
        
        if not supabase_url or not supabase_key:
            print("❌ Supabase 환경변수가 설정되지 않았습니다")
            print("   필요한 환경변수: SUPABASE_URL, SUPABASE_KEY")
            return {}
        
        # Supabase REST API 호출
        api_url = f"{supabase_url}/rest/v1/sku_total"
        headers = {
            "apikey": supabase_key,
            "Authorization": f"Bearer {supabase_key}",
            "Content-Type": "application/json"
        }
        
        # 품번코드와 상품명만 조회
        params = {
            "select": "품번코드,상품명"
        }
        
        response = requests.get(api_url, headers=headers, params=params, timeout=15)
        
        if response.status_code == 200:
            data = response.json()
            
            if not data:
                print("❌ Supabase에서 매핑 데이터가 없습니다")
                return {}
            
            # 매핑 딕셔너리 생성 {품번코드: 상품명}
            mapping = {}
            for row in data:
                product_code = str(row.get('품번코드', '')).strip()
                product_name = str(row.get('상품명', '')).strip()
                
                if product_code and product_name:
                    mapping[product_code] = product_name
            
            print(f"✅ Supabase 상품명 매핑 데이터 로드 완료: {len(mapping)}개")
            return mapping
            
        else:
            print(f"❌ Supabase API 호출 실패: {response.status_code}")
            print(f"❌ 응답: {response.text[:200]}")
            return {}
            
    except Exception as e:
        print(f"❌ Supabase 상품명 매핑 데이터 로드 실패: {e}")
        return {}

def update_orders_batch(order_ids, status, base_url, consumer_key, consumer_secret):
    """여러 주문 상태 배치 업데이트 (20개씩 안전하게 처리)"""
    import requests
    import json
    import time
    
    if not order_ids:
        return 0
    
    print(f"🔄 배치 업데이트 시작: {len(order_ids)}개 주문 → {status} 상태")
    
    # 20개씩 나누어 처리 (API 안정성)
    batch_size = 20
    total_updated = 0
    total_failed = 0
    
    for i in range(0, len(order_ids), batch_size):
        batch = order_ids[i:i+batch_size]
        batch_num = (i // batch_size) + 1
        total_batches = (len(order_ids) + batch_size - 1) // batch_size
        
        print(f"   📦 배치 {batch_num}/{total_batches}: {len(batch)}개 주문 처리 중...")
        
        # 배치 업데이트 데이터 구성
        batch_data = {
            "update": [
                {"id": int(order_id), "status": status} 
                for order_id in batch
            ]
        }
        
        try:
            # WooCommerce 배치 API 호출
            batch_url = f"{base_url}/wp-json/wc/v3/orders/batch"
            
            if base_url.startswith('https://'):
                params = {
                    'consumer_key': consumer_key,
                    'consumer_secret': consumer_secret
                }
                response = requests.post(
                    batch_url,
                    params=params,
                    headers={'Content-Type': 'application/json'},
                    data=json.dumps(batch_data),
                    timeout=30
                )
            else:
                auth = (consumer_key, consumer_secret)
                response = requests.post(
                    batch_url,
                    auth=auth,
                    headers={'Content-Type': 'application/json'},
                    data=json.dumps(batch_data),
                    timeout=30
                )
            
            if response.status_code == 200:
                result = response.json()
                updated_orders = result.get('update', [])
                
                success_count = 0
                for updated_order in updated_orders:
                    if updated_order.get('id'):
                        success_count += 1
                
                total_updated += success_count
                failed_count = len(batch) - success_count
                total_failed += failed_count
                
                print(f"   ✅ 배치 {batch_num} 완료: {success_count}개 성공, {failed_count}개 실패")
                
            else:
                print(f"   ❌ 배치 {batch_num} API 오류: {response.status_code}")
                print(f"   ❌ 응답: {response.text[:200]}")
                total_failed += len(batch)
                
        except Exception as e:
            print(f"   ❌ 배치 {batch_num} 처리 오류: {e}")
            total_failed += len(batch)
        
        # 배치 간 잠시 대기 (API 제한 방지)
        if i + batch_size < len(order_ids):
            time.sleep(0.5)
    
    print(f"🎉 배치 업데이트 완료: {total_updated}개 성공, {total_failed}개 실패")
    return total_updated

class ProcessingResults:
    """3PL 처리 결과 수집 클래스"""
    
    def __init__(self):
        self.reset()
    
    def reset(self):
        """결과 초기화"""
        self.domestic_orders = 0
        self.international_orders = 0
        self.digital_status_changes = 0
        self.reservation_status_changes = 0
        self.b2b_status_changes = 0
        self.happy_together_processed = 0
        self.errors = []
        self.warnings = []
    
    def add_domestic_orders(self, count):
        """국내 배송 주문 수 추가"""
        self.domestic_orders += count
    
    def add_international_orders(self, count):
        """국제 배송 주문 수 추가"""
        self.international_orders += count
    
    def add_digital_status_changes(self, count):
        """디지털 상품 상태 변경 수 추가"""
        self.digital_status_changes += count
    
    def add_reservation_status_changes(self, count):
        """예약 상품 상태 변경 수 추가"""
        self.reservation_status_changes += count
    
    def add_b2b_status_changes(self, count):
        """B2B 상품 상태 변경 수 추가"""
        self.b2b_status_changes += count
    
    def add_happy_together(self, count):
        """해피투게더 처리 수 추가"""
        self.happy_together_processed += count
    
    def add_error(self, error_msg):
        """오류 추가"""
        self.errors.append(error_msg)
    
    def add_warning(self, warning_msg):
        """경고 추가"""
        self.warnings.append(warning_msg)
    
    def add_korean_recipient_issue(self, korean_orders_df, site_name):
        """한글 수령인명으로 인한 EMS 제외 이슈 추가"""
        if korean_orders_df.empty:
            return
        
        issue_details = []
        issue_details.append(f"이상 발견: {site_name} 해외 발송 명단 - 수령인의 이름이 한글로 확인")
        issue_details.append(f"제외된 주문 {len(korean_orders_df)}개:")
        
        for idx, row in korean_orders_df.iterrows():
            order_id = row.get('주문번호', 'N/A')
            recipient = row.get('수령인명', 'N/A')
            address = str(row.get('배송지주소', ''))[:50] + "..." if len(str(row.get('배송지주소', ''))) > 50 else str(row.get('배송지주소', ''))
            issue_details.append(f"  • 주문번호 {order_id}: '{recipient}' → {address}")
        
        issue_details.append("※ EMS는 영문 수령인명만 허용되므로 해당 주문들은 주문서에서 제외되었습니다.")
        
        self.add_warning("\n".join(issue_details))
    
    def add_non_english_address_issue(self, non_english_orders_df, site_name):
        """비영문 주소로 인한 EMS 제외 이슈 추가"""
        if non_english_orders_df.empty:
            return
        
        issue_details = []
        issue_details.append(f"이상 발견: {site_name} 해외 발송 명단 - 주소에 비영문 문자 확인")
        issue_details.append(f"제외된 주문 {len(non_english_orders_df)}개:")
        
        for idx, row in non_english_orders_df.iterrows():
            order_id = row.get('주문번호', 'N/A')
            recipient = row.get('수령인명', 'N/A')
            address = str(row.get('배송지주소', ''))[:50] + "..." if len(str(row.get('배송지주소', ''))) > 50 else str(row.get('배송지주소', ''))
            issue_details.append(f"  • 주문번호 {order_id}: '{recipient}' → {address}")
        
        issue_details.append("※ EMS는 영문 주소만 허용되므로 해당 주문들은 관리자 확인이 필요합니다.")
        
        self.add_warning("\n".join(issue_details))
    
    def get_summary(self):
        """처리 결과 요약 반환"""
        summary = []
        summary.append("=== 3PL 처리 결과 요약 ===")
        summary.append("")
        summary.append("📊 처리된 주문:")
        summary.append(f"   🏠 국내 배송: {self.domestic_orders}건")
        summary.append(f"   🌍 국제 배송 (EMS): {self.international_orders}건")
        summary.append("")
        summary.append("🔄 상태 변경:")
        summary.append(f"   📱 디지털 상품 → shipped: {self.digital_status_changes}건")
        summary.append(f"   📦 예약 상품 → processing: {self.reservation_status_changes}건")
        summary.append(f"   🏢 B2B 상품 → shipped: {self.b2b_status_changes}건")
        summary.append("")
        summary.append("🎁 특수 처리:")
        summary.append(f"   👫 해피투게더 처리: {self.happy_together_processed}건")
        
        if self.errors:
            summary.append("")
            summary.append("❌ 오류 발생:")
            for error in self.errors:
                summary.append(f"   • {error}")
        
        if self.warnings:
            summary.append("")
            summary.append("⚠️ 주의사항:")
            for warning in self.warnings:
                summary.append(f"   • {warning}")
        
        if not self.errors and not self.warnings:
            summary.append("")
            summary.append("✅ 모든 처리가 성공적으로 완료되었습니다.")
        
        return "\n".join(summary)

# 전역 결과 수집기
processing_results = ProcessingResults()

def filter_po_box_orders(df):
    """사서함 주소가 포함된 주문을 분리하고 엑셀로 저장"""
    if df.empty:
        return df, None
    
    # 사서함 관련 키워드로 필터링 (대소문자 구분 없음, 정규식 이스케이프 처리)
    po_box_keywords = [
        '사서함',
        r'P\.O\.Box',     # P.O.Box (점을 이스케이프)
        r'P\.O\. Box',    # P.O. Box (점을 이스케이프)
        'PO Box',         # PO Box
        'POBox',          # POBox
        'po box'          # po box
    ]
    
    # 주소 컬럼에서 사서함 키워드 검색 (정규식 사용)
    po_box_pattern = '|'.join(po_box_keywords)
    po_box_mask = df['배송지주소'].str.contains(po_box_pattern, case=False, na=False, regex=True)
    
    # 사서함 주문 분리
    po_box_orders = df[po_box_mask].copy()
    regular_orders = df[~po_box_mask].copy()
    
    po_box_file_path = None
    
    if not po_box_orders.empty:
        print(f"📮 사서함 주소 주문 {len(po_box_orders)}개 발견")
        
        # 현재 날짜로 파일명 생성
        today_str = datetime.now().strftime('%y%m%d')
        po_box_file_path = os.path.join(DOWNLOAD_DIR, f"우체국용_사서함_주문_{today_str}.xlsx")
        
        try:
            # 엑셀 파일로 저장
            po_box_orders.to_excel(po_box_file_path, index=False, engine='openpyxl')
            print(f"📮 사서함 주문 저장 완료: {po_box_file_path}")
            
            # 처리 결과에 추가
            processing_results.add_warning(f"사서함 주소 주문 {len(po_box_orders)}개 별도 처리")
            
        except Exception as e:
            print(f"❌ 사서함 주문 파일 저장 실패: {e}")
            po_box_file_path = None
    else:
        print("✅ 사서함 주소 주문 없음")
    
    return regular_orders, po_box_file_path
